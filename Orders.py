import openpyxl as opxl
import traceback

class OrderReader:
    products = {}
#Read given file and send it to extract function
    def read_file(self,file):
        self.products.clear()
        wb = opxl.load_workbook(filename = file)
        sheet1 = wb.active
        self.products = self.extract(sheet1)
        return self.products
#Create new sheet - for writing results
    def new_sheet(self,wb,name):
        sheet = wb.create_sheet(name)
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 70
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 10
        sheet.column_dimensions["E"].width = 10
        return sheet
#Main write file funtion with results
    def write_file(self,file):
        wb = opxl.Workbook()
        cols = ["","Part Number","Nazwa produktu","Stan Obecny","Stan minimalny","Różnica","Sprzedaż","Czas do osiągnięcia minimalnego","Czas do wyprzedania"]
        sheet = self.new_sheet(wb,"Zamówienia")
        self.fill_sheet(sheet,self.products,cols)
        del wb['Sheet']
        wb.save(file)
#Fill active sheet with results
    def fill_sheet(self,sheet,dict,cols):
        columns = cols
        for x in range (1,len(cols)):
            sheet.cell(row=1,column=x).value = columns[x]
        i = 2
        for key in dict :
            values = dict[key]
            sheet.cell(row=i,column=1).value = key
            size = len(values)+2
            for col, val in enumerate(values,start=2):
                sheet.cell(row=i,column=col).value = val
            i += 1
#Function takes file path and takes data from crucial columns to create new dictionary
    def extract(self,sheet):
        temp_dict = {}
        shops = 7
        i=2
        while True:
            if sheet[f"A{i}"].value != None:
                #How list should look like
                #temp_list = ["name","stock","min","diff","sales","tomin","tonull"]
                if sheet[f"C{i}"].value not in temp_dict :
                    temp_dict[sheet[f"C{i}"].value] = [sheet[f"D{i}"].value,0,0,0,0,0,0]
                if sheet[f"A{i}"].value == 1:
                    temp_dict[sheet[f"C{i}"].value][4] = sheet[f"E{i}"].value
                elif sheet[f"A{i}"].value == 2:
                    temp_dict[sheet[f"C{i}"].value][1] = sheet[f"E{i}"].value
                elif sheet[f"A{i}"].value == 3:
                    temp_dict[sheet[f"C{i}"].value][2] = sheet[f"E{i}"].value * shops
                elif sheet[f"A{i}"].value == 4:
                    temp_dict[sheet[f"C{i}"].value][4] -= sheet[f"E{i}"].value
                temp_dict[sheet[f"C{i}"].value][3] = temp_dict[sheet[f"C{i}"].value][1] - temp_dict[sheet[f"C{i}"].value][2]
                if temp_dict[sheet[f"C{i}"].value][4] != 0:
                    temp_dict[sheet[f"C{i}"].value][5] = float("%.2f" %(temp_dict[sheet[f"C{i}"].value][3] / temp_dict[sheet[f"C{i}"].value][4]))
                    temp_dict[sheet[f"C{i}"].value][6] = float("%.2f" %(temp_dict[sheet[f"C{i}"].value][1] / temp_dict[sheet[f"C{i}"].value][4]))
                i += 1
            else :
                return temp_dict