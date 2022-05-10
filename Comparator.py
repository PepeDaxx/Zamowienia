import string
import openpyxl as opxl
import traceback
class Comparator:
    #Set up variables needed for process
    file1_dict = {}
    file2_dict = {}
    difference_dict = {}
    buy_difference_dict = {}
    new_items_dict = {}
    deleted_items_dict = {}
    names_dict = {}
    # With given .xlsx file extract only data that is crucial for calculations into dictionary
    def extract_to_dict(self,sheet):
        temp_dict = {}
        i = 17
        while True:
            if sheet[f"A{i}"].value != None:
                temp_dict[sheet[f"A{i}"].value] = [sheet[f"C{i}"].value,sheet[f"D{i}"].value,sheet[f"E{i}"].value]
                if f"A{i}" not in self.names_dict:
                    self.names_dict[sheet[f"A{i}"].value] =  sheet[f"B{i}"].value
                i += 1
            else :
                return temp_dict
    # With two given dicts (representing old and new buy prices) look for differences, promotions start and end, new items or items that doesn't exist in newer sheet
    def generate_results(self,dict1,dict2):
        for key in dict2:
            if key in dict1:
                if dict2[key][0] != dict1[key][0]:
                    difference = dict2[key][0] - dict1[key][0]
                    self.difference_dict[key] = [self.names_dict[key],"%.2f" % dict2[key][0],"%.2f" % dict1[key][0],"%.2f" % difference]
                if dict2[key][2] != dict1[key][2]:
                    new_price = dict2[key][1]
                    old_price = dict1[key][1]
                    if dict2[key][2] != None: 
                        new_price = dict2[key][2]
                        promotion = "Nowa"
                    if dict1[key][2] != None : 
                        old_price = dict1[key][2]
                        promotion = "Koniec"
                    if dict2[key][2] != None and dict1[key][2] != None:
                        if dict2[key][2] > dict1[key][2] : promotion = "Słabsza"
                        else : promotion = "Lepsza"
                    buy_difference = new_price - old_price
                    self.buy_difference_dict[key]= [self.names_dict[key],"%.2f" %new_price,"%.2f" %old_price,"%.2f" % buy_difference,promotion]
                elif dict2[key][1] != dict1[key][1] and dict2[key] == None:
                    buy_difference = dict2[key][1] - dict1[key][1]
                    self.buy_difference_dict[key]= [self.names_dict[key],"%.2f" %dict2[key][1],"%.2f" %dict1[key][1],"%.2f" % buy_difference,""]
            elif key not in dict1:
                price = dict2[key][1]
                promo =""
                if dict2[key][2] != None : promo="Tak"
                self.new_items_dict[key] = [self.names_dict[key],"%.2f" %dict2[key][0],"%.2f" %price,promo]
        for key in dict1:
            if key not in dict2:
                self.deleted_items_dict[key] = [self.names_dict[key]]
    # Function that reads .xlsx files from file paths and send it to extraction function to obtain two diffrent dictionaries (old,new)
    def read_files(self,file1=string,file2=string):
        try:
            self.file1_dict.clear()
            self.file2_dict.clear()
            wb1 = opxl.load_workbook(filename = file1)
            sheet1 = wb1['Cennik']
            wb2 = opxl.load_workbook(filename = file2)
            sheet2 = wb2['Cennik']
            print("Załadowane!")
            self.file1_dict = self.extract_to_dict(sheet1)
            self.file2_dict = self.extract_to_dict(sheet2)
            self.difference_dict.clear()
            self.new_items_dict.clear()
            self.generate_results(self.file1_dict,self.file2_dict)
            wb1.close()
            wb2.close()
            return self.difference_dict, self.new_items_dict, self.deleted_items_dict, self.buy_difference_dict
        except Exception as e:
            print(e)
    # Create new sheet with given details while writing a file with results
    def new_sheet(self,wb,name):
        sheet = wb.create_sheet(name)
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 70
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 10
        sheet.column_dimensions["E"].width = 10
        return sheet
    # Fill sheet with data while writing results
    def fill_sheet(self,sheet,dict,cols):
        columns = cols
        for x in range (1,len(cols)):
            sheet.cell(row=1,column=x).value = columns[x]
        i = 2
        for key in dict :
            values = dict[key]
            sheet.cell(row=i,column=1).value = key
            size = len(values)+2
            for col, val in enumerate(values,start=2):
                sheet.cell(row=i,column=col).value = val
            i += 1
    # Main function to write file with results of application
    def write_file(self,file_path):
        try:
            wb = opxl.Workbook()
            #Difference
            cols = ["","Part Number","Nazwa produktu","RRP","Stare RRP","Różnica","Promocja"]
            sheet = self.new_sheet(wb,"Różnice")
            self.fill_sheet(sheet,self.difference_dict,cols)
            cols.clear()
            #Buy Differences
            cols = ["","Part Number","Nazwa produktu","Cena Zakupu","Stara Cena","Różnica","Promocja"]
            sheet = self.new_sheet(wb,"Różnice Zakupowe")
            self.fill_sheet(sheet,self.buy_difference_dict,cols)
            cols.clear()
            #New
            cols = ["","Part Number","Nazwa produktu","RRP","Cena Zakupu","Promocja"]
            sheet = self.new_sheet(wb,"Nowe")
            self.fill_sheet(sheet,self.new_items_dict,cols)
            cols.clear()
            #Deleted
            cols = ["","Part Number","Nazwa produktu"]
            sheet = self.new_sheet(wb,"Usunięte")
            self.fill_sheet(sheet,self.deleted_items_dict,cols)
            cols.clear()
            #Close File
            del wb['Sheet']
            wb.save(file_path)
        except Exception as e:
            print(traceback.format_exc())